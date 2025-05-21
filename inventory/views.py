from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import LoginView
from django.contrib import messages
from django.db.models import Sum, F, Count, Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.template.loader import get_template
from django.core.paginator import Paginator
from django.db import transaction
from datetime import datetime, timedelta
import json
import csv
import io
from xhtml2pdf import pisa
import csv
from django.template.loader import get_template

from django.utils import timezone


from .models import (
    Category, Supplier, Product, Client, 
    Purchase, PurchaseItem, Sale, SaleItem, 
    Payment, StockMovement
)
from .forms import (
    LoginForm, CategoryForm, SupplierForm, ProductForm, 
    ClientForm, PurchaseForm, PurchaseItemForm, 
    SaleForm, SaleItemForm, PaymentForm,
    ClientPaymentForm, SupplierPaymentForm
)

class CustomLoginView(LoginView):
    form_class = LoginForm
    template_name = 'inventory/login.html'

@login_required
def dashboard(request):
    # Statistiques pour le tableau de bord
    total_products = Product.objects.count()
    low_stock_products = Product.objects.filter(quantity__lte=F('minimum_threshold')).count()
    total_suppliers = Supplier.objects.count()
    total_clients = Client.objects.count()
    
    # Ventes récentes (7 derniers jours)
    today = timezone.now()
    week_ago = today - timedelta(days=7)
    recent_sales = Sale.objects.filter(sale_date__gte=week_ago).order_by('-sale_date')[:5]
    
    # Produits les plus vendus
    top_products = SaleItem.objects.values('product__name').annotate(
        total_sold=Sum('quantity')
    ).order_by('-total_sold')[:5]
    
    # Achats récents
    recent_purchases = Purchase.objects.order_by('-purchase_date')[:5]
    
    # Données pour les graphiques
    # Ventes par jour (7 derniers jours)
    sales_by_day = []
    for i in range(7):
        day = today - timedelta(days=i)
        day_sales = Sale.objects.filter(
            sale_date__year=day.year,
            sale_date__month=day.month,
            sale_date__day=day.day
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        sales_by_day.append({
            'date': day.strftime('%d/%m'),
            'amount': float(day_sales)
        })
    sales_by_day.reverse()
    
    context = {
        'total_products': total_products,
        'low_stock_products': low_stock_products,
        'total_suppliers': total_suppliers,
        'total_clients': total_clients,
        'recent_sales': recent_sales,
        'top_products': top_products,
        'recent_purchases': recent_purchases,
        'sales_by_day': json.dumps(sales_by_day),
    }
    return render(request, 'inventory/dashboard.html', context)

# Gestion des catégories
@login_required
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'inventory/category/list.html', {'categories': categories})

@login_required
def category_create(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie créée avec succès.')
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'inventory/category/form.html', {'form': form})

@login_required
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie mise à jour avec succès.')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'inventory/category/form.html', {'form': form, 'category': category})

@login_required
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.success(request, 'Catégorie supprimée avec succès.')
        return redirect('category_list')
    return render(request, 'inventory/category/confirm_delete.html', {'category': category})

# Gestion des fournisseurs
@login_required
def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, 'inventory/supplier/list.html', {'suppliers': suppliers})

@login_required
def supplier_create(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fournisseur créé avec succès.')
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'inventory/supplier/form.html', {'form': form})

@login_required
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, 'Fournisseur mis à jour avec succès.')
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'inventory/supplier/form.html', {'form': form, 'supplier': supplier})

@login_required
def supplier_delete(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier.delete()
        messages.success(request, 'Fournisseur supprimé avec succès.')
        return redirect('supplier_list')
    return render(request, 'inventory/supplier/confirm_delete.html', {'supplier': supplier})

# Gestion des produits
@login_required
def product_list(request):
    products = Product.objects.all()
    categories = Category.objects.all()
    return render(request, 'inventory/product/list.html', {
        'products': products,
        'categories': categories
    })

@login_required
def product_create(request):
    if request.method == 'POST':
        form = ProductForm(request.POST)
        if form.is_valid():
            product = form.save()
            # Créer un mouvement de stock initial
            StockMovement.objects.create(
                product=product,
                movement_type='in',
                quantity=product.quantity,
                reference='Stock initial',
                notes='Création du produit',
                created_by=request.user
            )
            messages.success(request, 'Produit créé avec succès.')
            return redirect('product_list')
    else:
        form = ProductForm()
    return render(request, 'inventory/product/form.html', {'form': form})

@login_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    old_quantity = product.quantity
    
    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            new_product = form.save()
            new_quantity = new_product.quantity
            
            # Si la quantité a changé, créer un mouvement de stock
            if new_quantity != old_quantity:
                movement_type = 'in' if new_quantity > old_quantity else 'out'
                quantity = abs(new_quantity - old_quantity)
                
                StockMovement.objects.create(
                    product=new_product,
                    movement_type=movement_type,
                    quantity=quantity,
                    reference='Ajustement manuel',
                    notes=f'Ajustement de stock de {old_quantity} à {new_quantity}',
                    created_by=request.user
                )
            
            messages.success(request, 'Produit mis à jour avec succès.')
            return redirect('product_list')
    else:
        form = ProductForm(instance=product)
    
    return render(request, 'inventory/product/form.html', {'form': form, 'product': product})

@login_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        messages.success(request, 'Produit supprimé avec succès.')
        return redirect('product_list')
    return render(request, 'inventory/product/confirm_delete.html', {'product': product})

# Gestion des clients
@login_required
def client_list(request):
    clients = Client.objects.all()
    return render(request, 'inventory/client/list.html', {'clients': clients})

@login_required
def client_create(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Client créé avec succès.')
            return redirect('client_list')
    else:
        form = ClientForm()
    return render(request, 'inventory/client/form.html', {'form': form})

@login_required
def client_edit(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, 'Client mis à jour avec succès.')
            return redirect('client_list')
    else:
        form = ClientForm(instance=client)
    return render(request, 'inventory/client/form.html', {'form': form, 'client': client})

@login_required
def client_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        client.delete()
        messages.success(request, 'Client supprimé avec succès.')
        return redirect('client_list')
    return render(request, 'inventory/client/confirm_delete.html', {'client': client})

# Gestion des achats
@login_required
def purchase_list(request):
    purchases = Purchase.objects.all().order_by('-purchase_date')
    return render(request, 'inventory/purchase/list.html', {'purchases': purchases})

@login_required
@transaction.atomic
def purchase_create(request):
    if request.method == 'POST':
        form = PurchaseForm(request.POST)
        if form.is_valid():
            purchase = form.save(commit=False)
            purchase.created_by = request.user
            purchase.save()
            
            # Traiter les éléments d'achat depuis le formulaire JSON
            items_data = json.loads(request.POST.get('items_data', '[]'))
            total_amount = 0
            
            for item_data in items_data:
                product = Product.objects.get(pk=item_data['product_id'])
                quantity = int(item_data['quantity'])
                unit_price = float(item_data['unit_price'])
                
                # Créer l'élément d'achat
                purchase_item = PurchaseItem.objects.create(
                    purchase=purchase,
                    product=product,
                    quantity=quantity,
                    unit_price=unit_price
                )
                
                # Mettre à jour le stock du produit
                product.quantity += quantity
                product.save()
                
                # Créer un mouvement de stock
                StockMovement.objects.create(
                    product=product,
                    movement_type='in',
                    quantity=quantity,
                    reference=f'Achat #{purchase.id}',
                    purchase=purchase,
                    created_by=request.user
                )
                
                total_amount += quantity * unit_price
            
            # Mettre à jour le montant total de l'achat
            purchase.total_amount = total_amount
            purchase.save()
            
            messages.success(request, 'Achat créé avec succès.')
            return redirect('purchase_list')
    else:
        form = PurchaseForm()
    
    products = Product.objects.all()
    return render(request, 'inventory/purchase/form.html', {
        'form': form,
        'products': products
    })

@login_required
def purchase_detail(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    items = purchase.items.all()
    return render(request, 'inventory/purchase/detail.html', {
        'purchase': purchase,
        'items': items
    })

# Gestion des ventes
@login_required
def sale_list(request):
    sales = Sale.objects.all().order_by('-sale_date')
    return render(request, 'inventory/sale/list.html', {'sales': sales})

@login_required
@transaction.atomic
def sale_create(request):
    if request.method == 'POST':
        form = SaleForm(request.POST)
        if form.is_valid():
            sale = form.save(commit=False)
            sale.created_by = request.user
            sale.save()
            
            # Traiter les éléments de vente depuis le formulaire JSON
            items_data = json.loads(request.POST.get('items_data', '[]'))
            total_amount = 0
            
            for item_data in items_data:
                product = Product.objects.get(pk=item_data['product_id'])
                quantity = int(item_data['quantity'])
                unit_price = float(item_data['unit_price'])
                
                # Vérifier si le stock est suffisant
                if product.quantity < quantity:
                    messages.error(request, f'Stock insuffisant pour {product.name}. Disponible: {product.quantity}')
                    sale.delete()  # Annuler la vente
                    return redirect('sale_create')
                
                # Créer l'élément de vente
                sale_item = SaleItem.objects.create(
                    sale=sale,
                    product=product,
                    quantity=quantity,
                    unit_price=unit_price
                )
                
                # Mettre à jour le stock du produit
                product.quantity -= quantity
                product.save()
                
                # Créer un mouvement de stock
                StockMovement.objects.create(
                    product=product,
                    movement_type='out',
                    quantity=quantity,
                    reference=f'Vente #{sale.id}',
                    sale=sale,
                    created_by=request.user
                )
                
                total_amount += quantity * unit_price
            
            # Mettre à jour le montant total de la vente
            sale.total_amount = total_amount
            sale.save()
            
            messages.success(request, 'Vente créée avec succès.')
            return redirect('sale_list')
    else:
        form = SaleForm()
    
    products = Product.objects.all()
    clients = Client.objects.all()
    return render(request, 'inventory/sale/form.html', {
        'form': form,
        'products': products,
        'clients': clients
    })
    

@login_required
def sale_detail(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    items = sale.items.all()
    return render(request, 'inventory/sale/detail.html', {
        'sale': sale,
        'items': items
    })

@login_required
def generate_invoice_pdf(request, pk):
    sale = get_object_or_404(Sale, pk=pk)
    items = sale.items.all()
    
    template = get_template('inventory/sale/invoice_pdf.html')
    context = {
        'sale': sale,
        'items': items,
        'company_name': 'Votre Entreprise',
        'company_address': 'Adresse de votre entreprise',
        'company_phone': 'Téléphone de votre entreprise',
        'company_email': 'Email de votre entreprise',
    }
    html = template.render(context)
    
    # Créer un fichier PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="facture_{sale.id}.pdf"'
    
    # Générer le PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Une erreur est survenue lors de la génération du PDF.', status=500)
    
    return response


@login_required
def payment_list(request):
    payments = Payment.objects.all().order_by('-payment_date')
    
    # Calculer les totaux par type de paiement
    client_payments = payments.filter(payment_type='client')
    supplier_payments = payments.filter(payment_type='supplier')
    
    total_received = sum(payment.amount for payment in client_payments)
    total_paid = sum(payment.amount for payment in supplier_payments)
    
    context = {
        'payments': payments,
        'total_received': total_received,
        'total_paid': total_paid,
    }
    return render(request, 'inventory/payment/list.html', context)

@login_required
def client_payment_create(request):
    if request.method == 'POST':
        form = ClientPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.payment_type = 'client'
            payment.created_by = request.user
            
            # Récupérer la vente associée
            sale = form.cleaned_data['sale']
            payment.sale = sale
            payment.save()
            
            # Mettre à jour le montant payé de la vente
            sale.paid_amount += payment.amount
            sale.save()  # Le statut de paiement est mis à jour automatiquement dans le modèle
            
            messages.success(request, 'Paiement client enregistré avec succès.')
            return redirect('payment_list')
    else:
        form = ClientPaymentForm()
    
    return render(request, 'inventory/payment/client_form.html', {'form': form})


@login_required
def supplier_payment_create(request):
    if request.method == 'POST':
        form = SupplierPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.payment_type = 'supplier'
            payment.created_by = request.user
            
            # Récupérer l'achat associé
            purchase = form.cleaned_data['purchase']
            payment.purchase = purchase
            payment.save()
            
            # Mettre à jour le montant payé de l'achat
            purchase.paid_amount += payment.amount
            purchase.save()
            
            messages.success(request, 'Paiement fournisseur enregistré avec succès.')
            return redirect('payment_list')
    else:
        form = SupplierPaymentForm()
    
    return render(request, 'inventory/payment/supplier_form.html', {'form': form})

# Rapports et statistiques
@login_required
def stock_report(request):
    products = Product.objects.all().order_by('name')
    low_stock_products = products.filter(quantity__lte=F('minimum_threshold'))
    
    context = {
        'products': products,
        'low_stock_products': low_stock_products,
    }
    return render(request, 'inventory/reports/stock_report.html', context)

@login_required
def sales_report(request):
    # Filtres de date
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    sales = Sale.objects.all().order_by('-sale_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            
            sales = sales.filter(sale_date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, 'Format de date invalide.')
    
    # Calcul des totaux
    total_sales = sales.aggregate(total=Sum('total_amount'))['total'] or 0
    total_paid = sales.aggregate(total=Sum('paid_amount'))['total'] or 0
    total_due = total_sales - total_paid
    
    context = {
        'sales': sales,
        'total_sales': total_sales,
        'total_paid': total_paid,
        'total_due': total_due,
        'start_date': start_date_str,
        'end_date': end_date_str,
    }
    return render(request, 'inventory/reports/sales_report.html', context)

@login_required
def purchases_report(request):
    # Filtres de date
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    purchases = Purchase.objects.all().order_by('-purchase_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            
            purchases = purchases.filter(purchase_date__range=[start_date, end_date])
        except ValueError:
            messages.error(request, 'Format de date invalide.')
    
    # Calcul des totaux
    total_purchases = purchases.aggregate(total=Sum('total_amount'))['total'] or 0
    total_paid = purchases.aggregate(total=Sum('paid_amount'))['total'] or 0
    total_due = total_purchases - total_paid
    
    # Données pour les graphiques
    # 1. Tendance des achats (daily, weekly, monthly)
    purchases_trend_data = {
        'daily': [],
        'weekly': [],
        'monthly': []
    }
    
    # Données quotidiennes (30 derniers jours)
    end_date = timezone.now()
    start_date = end_date - timedelta(days=29)
    
    for i in range(30):
        date = start_date + timedelta(days=i)
        next_date = date + timedelta(days=1)
        
        daily_amount = Purchase.objects.filter(
            purchase_date__gte=date,
            purchase_date__lt=next_date
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        purchases_trend_data['daily'].append({
            'date': date.strftime('%d/%m'),
            'amount': float(daily_amount)
        })
    
    # Données hebdomadaires (12 dernières semaines)
    end_week = timezone.now()
    start_week = end_week - timedelta(weeks=11)
    
    for i in range(12):
        week_start = start_week + timedelta(weeks=i)
        week_end = week_start + timedelta(weeks=1)
        
        weekly_amount = Purchase.objects.filter(
            purchase_date__gte=week_start,
            purchase_date__lt=week_end
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        week_number = week_start.isocalendar()[1]
        purchases_trend_data['weekly'].append({
            'date': f'Sem {week_number}',
            'amount': float(weekly_amount)
        })
    
    # Données mensuelles (12 derniers mois)
    current_month = timezone.now().month
    current_year = timezone.now().year
    
    for i in range(12):
        month = ((current_month - i - 1) % 12) + 1
        year = current_year if month <= current_month else current_year - 1
        
        month_start = timezone.datetime(year, month, 1)
        if month == 12:
            month_end = timezone.datetime(year + 1, 1, 1)
        else:
            month_end = timezone.datetime(year, month + 1, 1)
        
        monthly_amount = Purchase.objects.filter(
            purchase_date__gte=month_start,
            purchase_date__lt=month_end
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        month_names = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin', 'Juil', 'Août', 'Sep', 'Oct', 'Nov', 'Déc']
        purchases_trend_data['monthly'].append({
            'date': f'{month_names[month-1]} {year}',
            'amount': float(monthly_amount)
        })
    
    # Inverser les listes pour avoir l'ordre chronologique
    purchases_trend_data['monthly'].reverse()
    
    # 2. Données des principaux fournisseurs
    top_suppliers = Supplier.objects.annotate(
        total_purchases=Sum('purchase__total_amount')
    ).exclude(total_purchases=None).order_by('-total_purchases')[:5]
    
    suppliers_data = []
    for supplier in top_suppliers:
        suppliers_data.append({
            'name': supplier.name,
            'purchases': float(supplier.total_purchases or 0)
        })
    
    # 3. Données des produits les plus achetés
    top_products = PurchaseItem.objects.values('product__name').annotate(
        total_quantity=Sum('quantity')
    ).order_by('-total_quantity')[:5]
    
    products_data = []
    for product in top_products:
        products_data.append({
            'name': product['product__name'],
            'quantity': product['total_quantity']
        })
    
    context = {
        'purchases': purchases,
        'total_purchases': total_purchases,
        'total_paid': total_paid,
        'total_due': total_due,
        'start_date': start_date_str,
        'end_date': end_date_str,
        'purchases_trend_data': json.dumps(purchases_trend_data),
        'suppliers_data': json.dumps(suppliers_data),
        'products_data': json.dumps(products_data)
    }
    return render(request, 'inventory/reports/purchases_report.html', context)

@login_required
def export_stock_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="stock_report.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Nom du produit', 'Catégorie', 'Prix d\'achat', 'Prix de vente', 'Quantité', 'Seuil minimum', 'Unité'])
    
    products = Product.objects.all().order_by('name')
    for product in products:
        writer.writerow([
            product.name,
            product.category.name if product.category else '',
            product.purchase_price,
            product.selling_price,
            product.quantity,
            product.minimum_threshold,
            product.unit
        ])
    
    return response

# API pour les opérations AJAX
@login_required
def get_product_info(request, product_id):
    try:
        product = Product.objects.get(pk=product_id)
        return JsonResponse({
            'id': product.id,
            'name': product.name,
            'purchase_price': float(product.purchase_price),
            'selling_price': float(product.selling_price),
            'quantity': product.quantity,
            'unit': product.unit
        })
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Produit non trouvé'}, status=404)



@login_required
def export_payments(request):
    # Récupérer les paramètres de la requête
    format_type = request.GET.get('format', 'pdf')
    include_client = request.GET.get('include_client') == '1'
    include_supplier = request.GET.get('include_supplier') == '1'
    
    # Filtrer les paiements selon les paramètres
    payments = Payment.objects.all().order_by('-payment_date')
    
    if include_client and not include_supplier:
        payments = payments.filter(payment_type='client')
    elif include_supplier and not include_client:
        payments = payments.filter(payment_type='supplier')
    
    # Exporter selon le format demandé
    if format_type == 'csv':
        # Créer un fichier CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="payments_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Date', 'Type', 'Méthode', 'Montant', 'Référence', 'Notes'])
        
        for payment in payments:
            reference = f"Vente #{payment.sale.id}" if payment.sale else (f"Achat #{payment.purchase.id}" if payment.purchase else "-")
            writer.writerow([
                payment.payment_date.strftime('%d/%m/%Y %H:%M'),
                payment.get_payment_type_display(),
                payment.get_payment_method_display(),
                payment.amount,
                reference,
                payment.notes or "-"
            ])
        
        return response
    
    elif format_type == 'excel':
        # Pour Excel, vous auriez besoin d'une bibliothèque comme openpyxl ou xlsxwriter
        # Voici un exemple simple avec openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapport des paiements"
            
            # En-têtes
            headers = ['Date', 'Type', 'Méthode', 'Montant', 'Référence', 'Notes']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Données
            for row_num, payment in enumerate(payments, 2):
                reference = f"Vente #{payment.sale.id}" if payment.sale else (f"Achat #{payment.purchase.id}" if payment.purchase else "-")
                ws.cell(row=row_num, column=1).value = payment.payment_date.strftime('%d/%m/%Y %H:%M')
                ws.cell(row=row_num, column=2).value = payment.get_payment_type_display()
                ws.cell(row=row_num, column=3).value = payment.get_payment_method_display()
                ws.cell(row=row_num, column=4).value = float(payment.amount)
                ws.cell(row=row_num, column=5).value = reference
                ws.cell(row=row_num, column=6).value = payment.notes or "-"
            
            # Ajuster la largeur des colonnes
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Créer la réponse HTTP
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="payments_report.xlsx"'
            
            # Sauvegarder le workbook dans la réponse
            wb.save(response)
            return response
            
        except ImportError:
            # Si openpyxl n'est pas installé, revenir au format CSV
            return HttpResponse("La bibliothèque Excel n'est pas disponible. Veuillez choisir un autre format.", status=400)
    
    else:  # PDF par défaut
        # Créer un PDF avec xhtml2pdf
        template = get_template('inventory/payment/payment_report_pdf.html')
        context = {
            'payments': payments,
            'total_received': sum(p.amount for p in payments if p.payment_type == 'client'),
            'total_paid': sum(p.amount for p in payments if p.payment_type == 'supplier'),
            'today': timezone.now(),
            'request': request,
        }
        html = template.render(context)
        
        # Créer un fichier PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="payments_report.pdf"'
        
        # Générer le PDF
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Une erreur est survenue lors de la génération du PDF.', status=500)
        
        return response



@login_required
def print_purchase(request, pk):
    purchase = get_object_or_404(Purchase, pk=pk)
    items = purchase.items.all()
    
    template = get_template('inventory/purchase/print.html')
    context = {
        'purchase': purchase,
        'items': items,
        'company_name': 'Votre Entreprise',
        'company_address': 'Adresse de votre entreprise',
        'company_phone': 'Téléphone de votre entreprise',
        'company_email': 'Email de votre entreprise',
    }
    html = template.render(context)
    
    # Créer un fichier PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bon_achat_{purchase.id}.pdf"'
    
    # Générer le PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Une erreur est survenue lors de la génération du PDF.', status=500)
    
    return response



@login_required
def export_purchases_report(request):
    # Récupérer les paramètres de la requête
    format_type = request.GET.get('format', 'pdf')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    include_charts = request.GET.get('include_charts') == '1'
    include_details = request.GET.get('include_details') == '1'
    
    # Filtrer les achats selon les dates
    purchases = Purchase.objects.all().order_by('-purchase_date')
    
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
            end_date = end_date.replace(hour=23, minute=59, second=59)
            
            purchases = purchases.filter(purchase_date__range=[start_date, end_date])
        except ValueError:
            pass
    
    # Calcul des totaux
    total_purchases = purchases.aggregate(total=Sum('total_amount'))['total'] or 0
    total_paid = purchases.aggregate(total=Sum('paid_amount'))['total'] or 0
    total_due = total_purchases - total_paid
    
    # Exporter selon le format demandé
    if format_type == 'csv':
        # Créer un fichier CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="purchases_report.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Référence', 'Date', 'Fournisseur', 'Montant Total', 'Montant Payé', 'Reste à Payer', 'Statut'])
        
        for purchase in purchases:
            status = "Payé" if purchase.paid_amount >= purchase.total_amount else "Partiellement payé" if purchase.paid_amount > 0 else "Non payé"
            writer.writerow([
                f"#{purchase.id}",
                purchase.purchase_date.strftime('%d/%m/%Y %H:%M'),
                purchase.supplier.name,
                purchase.total_amount,
                purchase.paid_amount,
                purchase.total_amount - purchase.paid_amount,
                status
            ])
        
        return response
    
    elif format_type == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.drawing.image import Image
            from io import BytesIO
            import matplotlib.pyplot as plt
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Rapport des achats"
            
            # Titre du rapport
            ws.merge_cells('A1:G1')
            cell = ws.cell(row=1, column=1)
            cell.value = "RAPPORT DES ACHATS"
            cell.font = Font(bold=True, size=16)
            cell.alignment = Alignment(horizontal='center')
            
            # Période
            ws.merge_cells('A2:G2')
            period_text = "Période: "
            if start_date_str and end_date_str:
                period_text += f"Du {start_date_str} au {end_date_str}"
            else:
                period_text += "Tous les achats"
            
            cell = ws.cell(row=2, column=1)
            cell.value = period_text
            cell.alignment = Alignment(horizontal='center')
            
            # Résumé
            ws.cell(row=4, column=1).value = "Total des achats:"
            ws.cell(row=4, column=2).value = float(total_purchases)
            ws.cell(row=4, column=2).number_format = '#,##0.00 "FCFA"'
            
            ws.cell(row=5, column=1).value = "Total payé:"
            ws.cell(row=5, column=2).value = float(total_paid)
            ws.cell(row=5, column=2).number_format = '#,##0.00 "FCFA"'
            
            ws.cell(row=6, column=1).value = "Reste à payer:"
            ws.cell(row=6, column=2).value = float(total_due)
            ws.cell(row=6, column=2).number_format = '#,##0.00 "FCFA"'
            
            # En-têtes du tableau
            headers = ['Référence', 'Date', 'Fournisseur', 'Montant Total', 'Montant Payé', 'Reste à Payer', 'Statut']
            header_row = 8
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Données
            for row_num, purchase in enumerate(purchases, header_row + 1):
                status = "Payé" if purchase.paid_amount >= purchase.total_amount else "Partiellement payé" if purchase.paid_amount > 0 else "Non payé"
                
                ws.cell(row=row_num, column=1).value = f"#{purchase.id}"
                ws.cell(row=row_num, column=2).value = purchase.purchase_date.strftime('%d/%m/%Y %H:%M')
                ws.cell(row=row_num, column=3).value = purchase.supplier.name
                ws.cell(row=row_num, column=4).value = float(purchase.total_amount)
                ws.cell(row=row_num, column=4).number_format = '#,##0.00 "FCFA"'
                ws.cell(row=row_num, column=5).value = float(purchase.paid_amount)
                ws.cell(row=row_num, column=5).number_format = '#,##0.00 "FCFA"'
                ws.cell(row=row_num, column=6).value = float(purchase.total_amount - purchase.paid_amount)
                ws.cell(row=row_num, column=6).number_format = '#,##0.00 "FCFA"'
                ws.cell(row=row_num, column=7).value = status
            
            # Ajuster la largeur des colonnes
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width
            
            # Ajouter des graphiques si demandé
            if include_charts:
                # Créer une nouvelle feuille pour les graphiques
                ws_charts = wb.create_sheet(title="Graphiques")
                
                                # Graphique de statut des paiements
                paid_count = purchases.filter(paid_amount__gte=F('total_amount')).count()
                partial_count = purchases.filter(paid_amount__gt=0).filter(paid_amount__lt=F('total_amount')).count()
                unpaid_count = purchases.filter(paid_amount=0).count()
                
                # Créer un graphique circulaire pour le statut des paiements
                if paid_count > 0 or partial_count > 0 or unpaid_count > 0:
                    plt.figure(figsize=(8, 6))
                    labels = ['Payé', 'Partiellement payé', 'Non payé']
                    sizes = [paid_count, partial_count, unpaid_count]
                    colors = ['#10B981', '#FBBF24', '#EF4444']
                    explode = (0.1, 0, 0)  # explode 1st slice
                    
                    plt.pie(sizes, explode=explode, labels=labels, colors=colors,
                            autopct='%1.1f%%', shadow=True, startangle=140)
                    plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle
                    plt.title('Statut des paiements')
                    
                    # Sauvegarder le graphique dans un buffer
                    img_buf = BytesIO()
                    plt.savefig(img_buf, format='png', dpi=100)
                    img_buf.seek(0)
                    plt.close()
                    
                    # Ajouter l'image au workbook
                    img = Image(img_buf)
                    img.width = 500
                    img.height = 375
                    ws_charts.add_image(img, 'A1')
                
                # Graphique des principaux fournisseurs
                top_suppliers = Supplier.objects.annotate(
                    total_purchases=Sum('purchases__total_amount')
                ).exclude(total_purchases=None).order_by('-total_purchases')[:5]
                
                if top_suppliers.exists():
                    plt.figure(figsize=(10, 6))
                    suppliers = [s.name for s in top_suppliers]
                    values = [float(s.total_purchases) for s in top_suppliers]
                    
                    plt.barh(suppliers, values, color='#0EA5E9')
                    plt.xlabel('Montant des achats (FCFA)')
                    plt.title('Principaux fournisseurs')
                    plt.grid(axis='x', linestyle='--', alpha=0.7)
                    
                    # Formater les valeurs sur l'axe x
                    plt.ticklabel_format(style='plain', axis='x')
                    
                    # Ajouter les valeurs à la fin de chaque barre
                    for i, v in enumerate(values):
                        plt.text(v + (max(values) * 0.01), i, f'{v:,.0f} FCFA', va='center')
                    
                    # Sauvegarder le graphique dans un buffer
                    img_buf = BytesIO()
                    plt.tight_layout()
                    plt.savefig(img_buf, format='png', dpi=100)
                    img_buf.seek(0)
                    plt.close()
                    
                    # Ajouter l'image au workbook
                    img = Image(img_buf)
                    img.width = 600
                    img.height = 350
                    ws_charts.add_image(img, 'A20')
            
            # Créer la réponse HTTP
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="purchases_report.xlsx"'
            
            # Sauvegarder le workbook dans la réponse
            wb.save(response)
            return response
            
        except ImportError:
            # Si openpyxl n'est pas installé, revenir au format CSV
            return HttpResponse("La bibliothèque Excel n'est pas disponible. Veuillez choisir un autre format.", status=400)
    
    else:  # PDF par défaut
        # Créer un PDF avec xhtml2pdf
        template = get_template('inventory/reports/purchases_report_pdf.html')
        
        # Données pour le graphique de statut des paiements
        paid_count = purchases.filter(paid_amount__gte=F('total_amount')).count()
        partial_count = purchases.filter(paid_amount__gt=0).filter(paid_amount__lt=F('total_amount')).count()
        unpaid_count = purchases.filter(paid_amount=0).count()
        
        # Données pour le graphique des principaux fournisseurs
        top_suppliers = Supplier.objects.annotate(
            total_purchases=Sum('purchase__total_amount')
        ).exclude(total_purchases=None).order_by('-total_purchases')[:5]
        
        suppliers_data = []
        for supplier in top_suppliers:
            suppliers_data.append({
                'name': supplier.name,
                'purchases': float(supplier.total_purchases or 0)
            })
        
        # Données pour le graphique des produits les plus achetés
        top_products = PurchaseItem.objects.values('product__name').annotate(
            total_quantity=Sum('quantity')
        ).order_by('-total_quantity')[:5]
        
        products_data = []
        for product in top_products:
            products_data.append({
                'name': product['product__name'],
                'quantity': product['total_quantity']
            })
        
        context = {
            'purchases': purchases if include_details else [],
            'total_purchases': total_purchases,
            'total_paid': total_paid,
            'total_due': total_due,
            'start_date': start_date_str,
            'end_date': end_date_str,
            'include_charts': include_charts,
            'payment_status': {
                'paid': paid_count,
                'partial': partial_count,
                'unpaid': unpaid_count
            },
            'suppliers_data': suppliers_data,
            'products_data': products_data,
            'today': timezone.now(),
            'company_name': 'BOCHEIRE',
            'company_address': 'Banikanni , Parakou',
            'company_phone': '+229 96 02 96 54',
            'company_email': 'contact@bocheire.com',
        }
        
        html = template.render(context)
        
        # Créer un fichier PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="purchases_report.pdf"'
        
        # Générer le PDF
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Une erreur est survenue lors de la génération du PDF.', status=500)
        
        return response
